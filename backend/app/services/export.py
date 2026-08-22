"""
Excel Export Service — Generate screening results spreadsheet.
Shows all candidates, which phase they were eliminated in, and their scores.
"""

import io
from datetime import datetime
from uuid import UUID
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from app.models.database_models import Candidate, JobDescription, Match


# Phase labels
PHASE_SIEVE_REJECTED = "Sieve (Pre-filter)"
PHASE_JUDGE_SCORED = "Judge (LLM Scored)"
PHASE_PASSED = "All Stages Passed"

# Color fills
FILL_HEADER = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
FILL_PASSED = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
FILL_JUDGE = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
FILL_SIEVE = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")

FONT_HEADER = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_NORMAL = Font(name="Calibri", size=10)
FONT_BOLD = Font(name="Calibri", bold=True, size=10)

THIN_BORDER = Border(
    left=Side(style="thin", color="cccccc"),
    right=Side(style="thin", color="cccccc"),
    top=Side(style="thin", color="cccccc"),
    bottom=Side(style="thin", color="cccccc"),
)


def _get_rejection_phase(match: Match) -> str:
    """Determine which phase a candidate was eliminated in."""
    if match.sieve_rank is not None and match.overall_score is None:
        # Was ranked by sieve but never scored by judge = sieve survivor that somehow wasn't judged
        # Or was filtered out by sieve
        return PHASE_SIEVE_REJECTED
    elif match.overall_score is not None:
        return PHASE_JUDGE_SCORED
    else:
        return PHASE_SIEVE_REJECTED


def _get_fill_for_phase(phase: str):
    """Return cell fill color based on phase."""
    if phase == PHASE_PASSED:
        return FILL_PASSED
    elif phase == PHASE_JUDGE_SCORED:
        return FILL_JUDGE
    else:
        return FILL_SIEVE


async def generate_screening_excel(
    db: AsyncSession,
    job_id: UUID,
) -> io.BytesIO:
    """
    Generate an Excel file with all resumes and their screening results.
    
    Columns:
    - Candidate Name
    - Email
    - Rejection Phase
    - Sieve Score
    - Sieve Rank
    - Overall Score
    - Skills Match
    - Experience Match
    - Education Match
    - Matched Skills
    - Missing Skills
    - Justification
    - Red Flags
    - Hallucination Flags
    - Blind Mode
    """
    # Load job
    job_result = await db.execute(select(JobDescription).where(JobDescription.id == job_id))
    job = job_result.scalar_one_or_none()
    if not job:
        raise ValueError(f"Job {job_id} not found")

    # Load all candidates
    candidates_result = await db.execute(select(Candidate))
    candidates = list(candidates_result.scalars().all())

    # Load matches for this job
    matches_result = await db.execute(
        select(Match).where(Match.job_id == job_id).options(selectinload(Match.candidate))
    )
    matches = list(matches_result.scalars().all())

    # Build lookup: candidate_id → match
    match_map = {m.candidate_id: m for m in matches}

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Screening Results"

    # ── Headers ──
    headers = [
        "Candidate Name",
        "Email",
        "Rejection Phase",
        "Sieve Score",
        "Sieve Rank",
        "Overall Score",
        "Skills Match",
        "Experience Match",
        "Education Match",
        "Matched Skills",
        "Missing Skills",
        "Justification",
        "Red Flags",
        "Hallucination Flags",
        "Blind Mode",
    ]

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Freeze header row
    ws.freeze_panes = "A2"

    # ── Data rows ──
    row_idx = 2
    for candidate in candidates:
        match = match_map.get(candidate.id)

        if match:
            # Determine phase
            if match.overall_score is not None:
                # Was scored by judge — check if it's a "good" score (passed threshold)
                # For this export, anyone scored by judge is "Judge Scored"
                phase = PHASE_JUDGE_SCORED
            else:
                phase = PHASE_SIEVE_REJECTED

            # Extract data
            sieve_score = match.sieve_score
            sieve_rank = match.sieve_rank
            overall_score = match.overall_score
            skills_match = match.skills_match
            experience_match = match.experience_match
            education_match = match.education_match

            judge_data = match.judge_json or {}
            matched_skills = ", ".join(judge_data.get("matched_skills", []))
            missing_skills = ", ".join(judge_data.get("missing_skills", []))
            justification = judge_data.get("justification", "")
            red_flags = ", ".join(judge_data.get("red_flags", []))

            grounding = match.grounding_flags or []
            hallucinations = [f for f in grounding if f.get("confidence") == "high"]
            hallucination_text = ", ".join([f.get("skill_claimed", "") for f in hallucinations])

            blind_mode = "Yes" if match.blind_mode else "No"
        else:
            # Candidate exists but wasn't part of this screening at all
            phase = "Not Screened"
            sieve_score = None
            sieve_rank = None
            overall_score = None
            skills_match = None
            experience_match = None
            education_match = None
            matched_skills = ""
            missing_skills = ""
            justification = ""
            red_flags = ""
            hallucination_text = ""
            blind_mode = "No"

        row_data = [
            candidate.name or "Unknown",
            candidate.email or "",
            phase,
            round(sieve_score, 4) if sieve_score is not None else "",
            sieve_rank if sieve_rank is not None else "",
            overall_score if overall_score is not None else "",
            skills_match if skills_match is not None else "",
            experience_match if experience_match is not None else "",
            education_match if education_match is not None else "",
            matched_skills,
            missing_skills,
            justification,
            red_flags,
            hallucination_text,
            blind_mode,
        ]

        # Determine row fill
        fill = _get_fill_for_phase(phase)

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = FONT_NORMAL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in [11, 12, 13, 14]))

            # Apply color only to phase column
            if col_idx == 3:
                cell.fill = fill
                cell.font = FONT_BOLD

        row_idx += 1

    # ── Summary Sheet ──
    ws_summary = wb.create_sheet("Summary")

    # Summary data
    total_candidates = len(candidates)
    sieve_survivors = sum(1 for m in matches if m.overall_score is not None)
    sieve_rejected = sum(1 for m in matches if m.overall_score is None and m.sieve_rank is not None)
    not_screened = total_candidates - len(matches)
    avg_score = 0
    if sieve_survivors > 0:
        avg_score = round(
            sum(m.overall_score for m in matches if m.overall_score is not None) / sieve_survivors, 1
        )

    summary_data = [
        ["Smart Screener — Screening Report", ""],
        ["", ""],
        ["Job Title", job.title],
        ["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["", ""],
        ["METRIC", "VALUE"],
        ["Total Candidates", total_candidates],
        ["Sieve Survivors (Sent to Judge)", sieve_survivors],
        ["Sieve Rejected", sieve_rejected],
        ["Not Screened", not_screened],
        ["Average Overall Score", f"{avg_score}/10"],
        ["Cost Reduction", f"{round((1 - sieve_survivors / max(total_candidates, 1)) * 100, 1)}%"],
    ]

    for row_idx, (label, value) in enumerate(summary_data, 1):
        cell_label = ws_summary.cell(row=row_idx, column=1, value=label)
        cell_value = ws_summary.cell(row=row_idx, column=2, value=value)

        if row_idx == 1:
            cell_label.font = Font(name="Calibri", bold=True, size=14)
        elif row_idx == 6:
            cell_label.font = FONT_HEADER
            cell_label.fill = FILL_HEADER
            cell_value.font = FONT_HEADER
            cell_value.fill = FILL_HEADER
        elif row_idx >= 7:
            cell_label.font = FONT_BOLD
            cell_value.font = FONT_NORMAL

    ws_summary.column_dimensions["A"].width = 35
    ws_summary.column_dimensions["B"].width = 30

    # ── Column widths for main sheet ──
    column_widths = [25, 30, 22, 12, 10, 12, 12, 15, 14, 35, 35, 50, 30, 30, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output
